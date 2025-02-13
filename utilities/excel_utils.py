import os
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
from PIL import Image
from openpyxl.drawing.image import Image as opImage
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium import webdriver

from models import Vehicle


def update_excel_with_new_data(vehicles: list[Vehicle], filename):
    """
    Update an existing Excel file with new data.
    If a vehicle already exists, then just adds new price to column.
    If a vehicle is new, then it adds a new row.
    """

    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    last_column = worksheet.max_column + 1
    # Add new column for new prices
    worksheet.cell(column=last_column, row=1).value = f"Price {date.today().strftime('%d-%m-%Y')}"

    for vehicle in vehicles:
        found = False
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=2, min_col=2):
            if found:
                break
            for cell in row:  # this is dumb, but how do you get the single cell from that row??
                if cell.value == vehicle.id:
                    worksheet.cell(cell.row, last_column).value = vehicle.price
                    found = True
                    break
        if not found:
            # add new line, since vehicle is not in the excel
            worksheet.append([i for i in vehicle.to_dict().values()])
            worksheet.cell(worksheet.max_row, last_column).value = vehicle.price
            print(f"Added new vehicle to row {worksheet.max_row}: {vehicle.make} {vehicle.model}. "
                  f"Price: {vehicle.price} || Link: {vehicle.link}")

    print(f"Added new prices to column {last_column}")
    # Add styling to prices, links and headers
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=last_column, max_col=last_column):
        for cell in row:
            cell.number_format = '#0€'

    make_link_column_hyperlinks(worksheet)
    update_header_styles(worksheet)

    # Save the workbook
    workbook.save(filename)


def make_link_column_hyperlinks(worksheet):
    """
    Link column: F
    """
    for cell in worksheet['F']:
        cell.hyperlink = cell.value
        cell.style = 'Hyperlink'


def save_vehicles_to_excel(vehicles, filename, images: bool = True):
    """
    TODO: make this not do selenium stuff, clean this up......................
    """
    df = pd.DataFrame([vehicle.to_dict_with_price() for vehicle in vehicles])

    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Vehicles"

    # Write the DataFrame to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

    # Add style to price and link cells
    for cell in worksheet['M']:
        cell.number_format = '#0€'  # Set the price column to display as currency

    make_link_column_hyperlinks(worksheet)

    update_header_styles(worksheet)

    if images:
        driver = webdriver.Chrome()
        thumb_path = os.path.join(os.getcwd(), "thumbs")
        # Insert images into the worksheet
        for idx, vehicle in enumerate(vehicles, start=2):  # start=2 to account for header row
            img_path = os.path.join(thumb_path, f"{vehicle.id}.png")
            if Path(img_path).is_file():
                continue
            driver.get(vehicle.image_url)
            img_element = driver.find_element("tag name", "img")
            img_element.screenshot(img_path)
            pil_image = Image.open(img_path)
            pil_image.close()
            img_openpyxl = opImage(img_path)
            worksheet.add_image(img_openpyxl, f'H{idx}')  # Assuming the image column is H
            worksheet.row_dimensions[idx].height = 200  # Set the row height to accommodate the image

    # Save the workbook
    workbook.save(filename)


def update_header_styles(worksheet):
    # Add bottom borders and style to header cells
    border = Border(bottom=Side(border_style="medium"))
    for i in range(1, worksheet.max_column + 1):
        worksheet.cell(row=1, column=i).border = border
        worksheet.cell(row=1, column=i).style = 'Pandas'
